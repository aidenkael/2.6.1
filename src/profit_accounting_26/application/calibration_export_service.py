"""校准反馈导出 V2（阶段 3 升级）。

输出普通目录（不是 ZIP）：

``校准反馈_YYYYMMDD_HHMMSS/``
├─ ``校准反馈.xlsx``
├─ ``manifest.json``
└─ ``images/``

- Sheet1 ``校准反馈`` 严格 7 列，只回答“AI 第一次怎么判断 → 用户后来怎么修正 →
  真实头程是什么”，禁止经济字段与 current_estimate；
- Sheet1 ``图片`` 列直接嵌入每条记录第一张主图的缩略图（不嵌高清原图）；
- ``manifest.json`` 供 Agent 直接读取结构化校准数据，字段语义与 Sheet1 7 列一致，
  每条 record 新增 ``machine_facts``（ai_initial / user_feedback 两层机器可读事实），
  UTF-8 / ensure_ascii=False，图片全部相对路径，不含任何经济字段与 current_estimate；
- Sheet2 ``导出信息`` 只放技术元数据（record_id / batch / exported_at /
  image_relative_paths / contract_version）；
- 导出状态按 ``record_id`` 保存在 ``data_dir/calibration/export_state.json``，
  只有 preflight、图片复制、Excel、manifest、状态落盘全部真正成功后才标记已导出；
- 图片优先原图，原图缺失时允许 thumbnail fallback 并记录 warning；
  记录声明有图但原图与缩略图都缺失时批次失败，不静默成功。
"""

from __future__ import annotations

import io
import json
import shutil
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from profit_accounting_26.application.calibration_feedback_service import CalibrationFeedbackService

CONTRACT_VERSION = "Calibration Feedback Export V2"
EXPORT_STATE_FILE = "export_state.json"
MANIFEST_FILE = "manifest.json"
SHEET1_COLUMNS = (
    "序号",
    "商品简名",
    "商品链接",
    "图片",
    "AI首次发货估算",
    "用户校准内容",
    "真实头程",
)
SHEET2_COLUMNS = (
    "record_id",
    "export_batch_id",
    "exported_at",
    "image_relative_paths",
    "contract_version",
)
MODES = ("all", "range", "pending")

# ---- V2 machine_facts 白名单：包装分析真正需要的字段，其余默认不进 manifest ----
_OBSERVATION_FIELDS = frozenset({
    "product_name",
    "product_type",
    "product_family",
    "product_type_raw",
    "product_type_code",
    "product_family_code",
    "material",
    "material_family",
    "material_family_code",
    "packaging_state_hint",
    "display_product_summary",
    "display_packaging_summary",
    "overall_form",
    "packing_actions",
    "packing_constraints",
    "rigidity",
    "foldability",
    "compressibility",
    "requires_shape_retention",
    "has_hard_bottom",
    "has_hard_backboard",
    "has_frame",
    "has_rigid_insert",
    "has_rigid_parts",
    "retail_box_visible",
    "hard_card_visible",
    "protrusion_flattenable",
    "length_cm",
    "width_cm",
    "height_cm",
    "weight_g",
    "dimension_value_source",
    "weight_value_source",
    "dimension_scope",
    "weight_scope",
    "quantity",
    "quantity_source",
    "quantity_summary",
    "source",
    "model",
    "prompt_version",
    "confidence",
})

_EVIDENCE_FIELDS = frozenset({
    "product_type",
    "product_family",
    "material",
    "rigidity",
    "foldability",
    "compressibility",
    "requires_shape_retention",
    "has_hard_bottom",
    "has_hard_backboard",
    "has_frame",
    "has_rigid_insert",
    "has_rigid_parts",
    "retail_box_visible",
    "hard_card_visible",
    "protrusion_flattenable",
    "length_cm",
    "width_cm",
    "height_cm",
    "weight_g",
    "dimension_scope",
    "weight_scope",
    "overall_form",
    "packing_actions",
    "packing_constraints",
    "quantity",
})

_SHIPMENT_FIELDS = frozenset({
    "length_cm", "width_cm", "height_cm", "weight_g", "state", "packaging_method",
})

_PACKAGING_SCENARIO_FIELDS = frozenset({
    "packaging_state",
    "packaging_method",
    "length_cm",
    "width_cm",
    "height_cm",
    "weight_g",
    "reasoning_summary",
    "confidence",
    "needs_review",
    "default_fields_used",
})

# 递归防护：manifest 任何层级的键名命中以下子串即整体剔除（含嵌套 raw evidence）
_FORBIDDEN_KEY_PARTS = (
    "current_estimate",
    "calculation_snapshot",
    "profit_scenarios",
    "product_cost",
    "domestic_shipping",
    "system_cost",
    "exchange_rate",
    "sale_price",
    "profit",
    "subsidy",
    "tail_fee",
    "shein_quote",
)

# Excel 嵌入主图缩略图：只嵌第一张，尺寸限制在 100～140px 保持清晰
_EXCEL_MAIN_IMAGE_MAX_PX = 112
_EXCEL_ROW_HEIGHT_PT = 88


def _utc_now_iso() -> str:
    return datetime.now(UTC).isoformat()


def _num(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _fmt_dims_weight(length: Any, width: Any, height: Any, weight: Any) -> str:
    length = _num(length)
    width = _num(width)
    height = _num(height)
    weight = _num(weight)
    if length is None and width is None and height is None and weight is None:
        return ""
    dims = "×".join(
        "—" if value is None else f"{value:g}"
        for value in (length, width, height)
    )
    if weight is not None:
        return f"{dims} cm / {weight:g}g"
    return f"{dims} cm"


class ExportStateStore:
    """轻量独立导出状态（JSON），按 record_id 记录。"""

    def __init__(self, data_dir: str | Path) -> None:
        self.path = Path(data_dir) / "calibration" / EXPORT_STATE_FILE

    def load(self) -> dict[str, Any]:
        if not self.path.is_file():
            return {"records": {}}
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError):
            return {"records": {}}
        records = data.get("records")
        return {"records": records if isinstance(records, dict) else {}}

    def exported_record_ids(self) -> set[str]:
        return {str(record_id) for record_id in self.load()["records"]}

    def mark_exported(self, record_ids: list[str], *, batch_id: str, exported_at: str) -> None:
        """只有整个批次真正成功后调用；取消/异常不得调用。"""
        state = self.load()
        for record_id in record_ids:
            state["records"][str(record_id)] = {
                "export_batch_id": batch_id,
                "exported_at": exported_at,
            }
        self.path.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.path.with_suffix(self.path.suffix + ".tmp")
        temporary.write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        temporary.replace(self.path)


@dataclass(slots=True)
class ExportResult:
    output_dir: Path
    batch_id: str
    exported_at: str
    record_ids: list[str]
    warnings: list[str] = field(default_factory=list)


class ExportIncompleteError(ValueError):
    """批次不完整（图片缺失等）：不得标记已导出。"""


def _image_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    items = payload.get("images") if isinstance(payload.get("images"), list) else []
    return [item for item in items if isinstance(item, dict)]


def _ai_initial_block(payload: dict[str, Any]) -> dict[str, Any]:
    v2 = payload.get("_v2") if isinstance(payload.get("_v2"), dict) else {}
    initial = v2.get("ai_initial")
    return initial if isinstance(initial, dict) else {}


def _pick_fields(source: dict[str, Any], allowed: frozenset[str]) -> dict[str, Any]:
    return {key: value for key, value in source.items() if key in allowed}


def _is_forbidden_key(key: str) -> bool:
    lowered = str(key).lower()
    return any(part in lowered for part in _FORBIDDEN_KEY_PARTS)


def _strip_forbidden_keys(value: Any) -> Any:
    """递归剔除键名命中经济字段子串的条目（manifest 最终防线）。"""
    if isinstance(value, dict):
        return {
            key: _strip_forbidden_keys(item)
            for key, item in value.items()
            if not _is_forbidden_key(key)
        }
    if isinstance(value, list):
        return [_strip_forbidden_keys(item) for item in value]
    return value


def _evidence_block(observation: dict[str, Any]) -> dict[str, Any]:
    """raw_payload → 过滤后的 evidence：只保留与包装有关的原始证据。

    不允许整体导出 raw_payload；field_evidence / confirmed_facts 走包装白名单，
    shipment 只保留 6 个字段；经济字段即使嵌套也由递归剔除兜底。
    """
    raw_payload = observation.get("raw_payload")
    raw_payload = raw_payload if isinstance(raw_payload, dict) else {}
    evidence: dict[str, Any] = {}
    for key in ("field_evidence", "confirmed_facts"):
        source = raw_payload.get(key)
        if isinstance(source, dict):
            evidence[key] = _pick_fields(source, _EVIDENCE_FIELDS)
    issue = raw_payload.get("dimension_semantic_issue")
    if isinstance(issue, str) and issue.strip():
        evidence["dimension_semantic_issue"] = issue
    shipment = raw_payload.get("shipment")
    if isinstance(shipment, dict):
        evidence["shipment"] = _pick_fields(shipment, _SHIPMENT_FIELDS)
    return _strip_forbidden_keys(evidence)


def _packaging_scenario_block(scenario: dict[str, Any]) -> dict[str, Any]:
    return {key: scenario.get(key) for key in _PACKAGING_SCENARIO_FIELDS}


def _packaging_proposal_block(initial: dict[str, Any]) -> dict[str, Any] | None:
    """AI 首次包装方案：优先 external_ai_packaging_proposal，回退 adopted_packaging。"""
    source_kind = "external_ai_packaging_proposal"
    proposal = initial.get("external_ai_packaging_proposal")
    if not isinstance(proposal, dict):
        adopted = initial.get("adopted_packaging")
        if isinstance(adopted, dict):
            proposal = adopted
            source_kind = "ai_initial.adopted_packaging"
    if not isinstance(proposal, dict):
        return None
    normal = proposal.get("normal")
    conservative = proposal.get("conservative")
    return {
        "source_kind": source_kind,
        "normal": _packaging_scenario_block(normal) if isinstance(normal, dict) else None,
        "conservative": (
            _packaging_scenario_block(conservative) if isinstance(conservative, dict) else None
        ),
        "engine_version": proposal.get("engine_version") or initial.get("engine_version"),
        "calibration_version": proposal.get("calibration_version") or initial.get("calibration_version"),
    }


def _machine_ai_initial(payload: dict[str, Any]) -> dict[str, Any] | None:
    """machine_facts.ai_initial：只从真正首次 AI 快照（_v2.ai_initial）取包装事实。

    旧记录（无 _v2.ai_initial 或只有 legacy_layers_ai_raw）返回 None，不伪造精确事实。
    """
    initial = _ai_initial_block(payload)
    if not initial or "legacy_layers_ai_raw" in initial:
        return None
    block: dict[str, Any] = {}
    for key in ("provider", "model", "prompt_version", "engine_version", "calibration_version"):
        value = initial.get(key)
        if value not in (None, ""):
            block[key] = value
    observation = initial.get("observation")
    if isinstance(observation, dict):
        filtered = _strip_forbidden_keys(_pick_fields(observation, _OBSERVATION_FIELDS))
        if filtered:
            block["observation"] = filtered
        evidence = _evidence_block(observation)
        if evidence:
            block["evidence"] = evidence
    proposal = _packaging_proposal_block(initial)
    if proposal is not None:
        block["packaging_proposal"] = proposal
    return block or None


def _machine_local_adopted(initial: dict[str, Any]) -> dict[str, Any] | None:
    """machine_facts.local_adopted：第一次本地仲裁后的最终采用结果。

    只从 ai_initial.adopted_packaging 读取（AI 首次识图时的本地仲裁输出），
    保存尺寸/重量、proposal_source、applied rule ids、conflicts、adjustments
    与 engine/calibration version，供 Logistics-calibration 判断
    “AI 错了？本地规则改错了？用户后来改了？”。
    """
    adopted = initial.get("adopted_packaging") if isinstance(initial, dict) else None
    if not isinstance(adopted, dict):
        return None
    conservative = adopted.get("conservative")
    normal = adopted.get("normal")
    tier = conservative if isinstance(conservative, dict) else normal
    if not isinstance(tier, dict):
        return None
    block: dict[str, Any] = {}
    shipment = _packaging_scenario_block(tier)
    if shipment:
        block["shipment"] = shipment
    for key in ("proposal_source", "engine_version", "calibration_version"):
        value = adopted.get(key)
        if value not in (None, ""):
            block[key] = value
    applied = adopted.get("applied_profile_ids")
    if isinstance(applied, list):
        block["applied_rule_ids"] = [str(rule_id) for rule_id in applied if str(rule_id)]
    conflicts = adopted.get("conflicts")
    if isinstance(conflicts, list):
        block["conflicts"] = list(conflicts)
    adjustments = adopted.get("adjustments")
    if isinstance(adjustments, list):
        block["adjustments"] = list(adjustments)
    return block or None


def _machine_user_feedback(feedback) -> dict[str, Any] | None:
    """machine_facts.user_feedback：直接读取 linked feedback 的精确字段。

    只导出规则分析事实，不含 calibration_exported_at / export_batch_id 等
    软件内部导出状态；实际费用与真实包装尺寸保持各自原值，不互相推导。
    """
    if feedback is None:
        return None
    return {
        "feedback_id": feedback.feedback_id,
        "feedback_schema_version": feedback.feedback_schema_version,
        "source": feedback.source,
        "created_at": feedback.created_at,
        "updated_at": feedback.updated_at,
        "structure": feedback.structure.to_dict(),
        "suggested_package": (
            feedback.suggested_package.to_dict() if feedback.suggested_package else None
        ),
        "actual_logistics": (
            feedback.actual_logistics.to_dict() if feedback.actual_logistics else None
        ),
        "user_note": feedback.user_note,
    }


def first_ai_short_name(payload: dict[str, Any]) -> str:
    """AI 第一次识别产生的简化商品名；没有真正的首次 AI 简名时返回空串。"""
    initial = _ai_initial_block(payload)
    observation = initial.get("observation")
    if not isinstance(observation, dict):
        return ""
    name = str(observation.get("product_name") or "").strip()
    if name:
        return name
    return str(observation.get("display_product_summary") or "").strip()


def first_ai_shipment_text(payload: dict[str, Any]) -> str:
    """AI 首次发货估算：尺寸/重量 + 发货判断，一行一个信息块。"""
    initial = _ai_initial_block(payload)
    if not initial:
        return ""
    observation = initial.get("observation")
    observation = observation if isinstance(observation, dict) else {}
    raw_payload = observation.get("raw_payload")
    raw_payload = raw_payload if isinstance(raw_payload, dict) else {}
    shipment = raw_payload.get("shipment")
    shipment = shipment if isinstance(shipment, dict) else {}

    normal: dict[str, Any] = {}
    proposal = initial.get("external_ai_packaging_proposal")
    proposal = proposal if isinstance(proposal, dict) else {}
    candidate = proposal.get("normal")
    if isinstance(candidate, dict):
        normal = candidate
    if not normal:
        adopted = initial.get("adopted_packaging")
        adopted = adopted if isinstance(adopted, dict) else {}
        candidate = adopted.get("normal")
        if isinstance(candidate, dict):
            normal = candidate

    length = normal.get("length_cm", shipment.get("length_cm"))
    width = normal.get("width_cm", shipment.get("width_cm"))
    height = normal.get("height_cm", shipment.get("height_cm"))
    weight = normal.get("weight_g", shipment.get("weight_g"))
    dims = _fmt_dims_weight(length, width, height, weight)
    state = str(shipment.get("state") or normal.get("packaging_method") or "").strip()
    if not dims and not state:
        return ""
    lines = [part for part in (dims, state) if part]
    return "\n".join(lines)


def user_calibration_text(feedback) -> str:
    """只读真正用户层：user_note + 真正 user_suggested 的 suggested_package。"""
    lines: list[str] = []
    note = str(getattr(feedback, "user_note", "") or "").strip().replace("\n", " ")
    if note:
        lines.append(f"用户反馈：{note}")
    suggested = getattr(feedback, "suggested_package", None)
    if suggested is not None and suggested.has_content():
        text = _fmt_dims_weight(
            suggested.length_cm, suggested.width_cm, suggested.height_cm, suggested.weight_g
        )
        method = str(suggested.packaging_method or "").strip()
        if method:
            text = f"{text}；{method}" if text else method
        lines.append(f"建议包装：{text}")
    return "\n".join(lines)


def actual_first_mile_text(feedback) -> str:
    """真实头程：只读 actual_forwarder + actual_first_mile_fee_rmb。"""
    actual = getattr(feedback, "actual_logistics", None)
    if actual is None or actual.actual_first_mile_fee_rmb is None:
        return ""
    forwarder = str(actual.actual_forwarder or "").strip() or "—"
    return f"{forwarder} / ¥{actual.actual_first_mile_fee_rmb:.2f}"


def parse_seq_range(text: str, record_count: int) -> tuple[int, int]:
    """解析 1-30 形式的连续序号范围（1 基、含端点），非法即抛错。"""
    raw = str(text or "").strip()
    if not raw:
        raise ValueError("自定义范围不能为空，例如 1-30")
    parts = raw.split("-")
    if len(parts) != 2:
        raise ValueError(f"范围格式错误：{raw}（应为 1-30）")
    try:
        start, end = int(parts[0]), int(parts[1])
    except ValueError as exc:
        raise ValueError(f"范围必须是数字：{raw}") from exc
    if start < 1 or end < start:
        raise ValueError(f"范围非法：{raw}（开始必须 ≥1 且开始 ≤ 结束）")
    if end > record_count:
        raise ValueError(f"范围超出当前记录数：{raw}（当前共 {record_count} 条）")
    return start, end


class CalibrationFeedbackExporter:
    """校准反馈导出 V2 执行器：模式选择 → preflight → 复制图片 → 写 Excel → 标记状态。"""

    def __init__(
        self,
        *,
        data_dir: str | Path,
        feedback_service: CalibrationFeedbackService,
        state_store: ExportStateStore | None = None,
    ) -> None:
        self.data_dir = Path(data_dir)
        self.feedback_service = feedback_service
        self.state_store = state_store or ExportStateStore(self.data_dir)

    # ------------------------------------------------------------------ 图片

    def _image_source(self, item: dict[str, Any]) -> tuple[Path, str, bool]:
        """返回 (源文件路径, fallback 标识)。原图优先，缩略图 fallback。"""
        storage_key = item.get("storage_key") or item.get("relative_path")
        thumbnail_key = item.get("thumbnail_key")
        if storage_key:
            original = self.data_dir / str(storage_key)
            if original.is_file():
                return original, "original", False
        if thumbnail_key:
            thumbnail = self.data_dir / str(thumbnail_key)
            if thumbnail.is_file():
                return thumbnail, "thumbnail", True
        raise ExportIncompleteError(
            f"记录 {item.get('image_id') or item.get('sha256') or '?'} 的图片缺失："
            "原图与缩略图均不存在"
        )

    # ------------------------------------------------------------------ 导出

    def export(
        self,
        records: list[dict[str, Any]],
        mode: str,
        target_parent: str | Path,
        *,
        seq_range: str | None = None,
    ) -> ExportResult:
        if mode not in MODES:
            raise ValueError(f"未知导出模式: {mode}")
        if not records:
            raise ValueError("当前没有可导出的历史记录")

        # 1) 选择本批记录（保持页面显示顺序；序号 → record_id 映射只在本页有效）
        if mode == "range":
            start, end = parse_seq_range(seq_range, len(records))
            selected = records[start - 1:end]
        elif mode == "pending":
            exported = self.state_store.exported_record_ids()
            selected = [payload for payload in records if str(payload.get("id") or "") not in exported]
        else:
            selected = list(records)
        if not selected:
            raise ValueError("当前没有可导出的记录（未导出部分为空）")

        record_ids = [str(payload.get("id") or "") for payload in selected]

        # 2) preflight：图片缺失在写文件前发现，避免写一半才发现大量缺失
        image_plan: list[list[tuple[Path, bool]]] = []
        for payload in selected:
            per_record: list[tuple[Path, bool]] = []
            for item in _image_items(payload):
                source, _kind, is_fallback = self._image_source(item)
                per_record.append((source, is_fallback))
            image_plan.append(per_record)

        # 3) 创建输出目录并复制图片
        now = datetime.now(UTC)
        batch_id = uuid4().hex
        exported_at = now.isoformat()
        base_name = f"校准反馈_{now:%Y%m%d_%H%M%S}"
        output_dir = Path(target_parent) / base_name
        counter = 1
        while output_dir.exists():
            output_dir = Path(target_parent) / f"{base_name}-{counter}"
            counter += 1
        images_dir = output_dir / "images"
        images_dir.mkdir(parents=True, exist_ok=False)

        warnings: list[str] = []
        image_relative_paths: list[list[str]] = []
        try:
            for seq, (payload, per_record) in enumerate(zip(selected, image_plan, strict=False), start=1):
                relative_paths: list[str] = []
                for index, (source, is_fallback) in enumerate(per_record, start=1):
                    suffix = source.suffix.lower() or ".jpg"
                    target_name = f"{seq:03d}_{index}{suffix}"
                    shutil.copy2(source, images_dir / target_name)
                    relative_paths.append(f"images/{target_name}")
                    if is_fallback:
                        warnings.append(
                            f"记录 {payload.get('id')} 第 {index} 张图原图缺失，已使用缩略图 fallback"
                        )
                image_relative_paths.append(relative_paths)
        except OSError as exc:
            raise ExportIncompleteError(f"图片复制失败：{exc}") from exc

        # 4) 写 Excel（写入失败 = 批次失败，不标记）
        excel_path = output_dir / "校准反馈.xlsx"
        try:
            self._write_excel(
                excel_path,
                records=selected,
                image_relative_paths=image_relative_paths,
                batch_id=batch_id,
                exported_at=exported_at,
            )
        except OSError as exc:
            raise ExportIncompleteError(f"Excel 写入失败：{exc}") from exc

        # 5) 写 manifest.json（供 Agent 直接读取；写失败同样不得显示导出完成）
        manifest_path = output_dir / MANIFEST_FILE
        try:
            self._write_manifest(
                manifest_path,
                records=selected,
                image_relative_paths=image_relative_paths,
                batch_id=batch_id,
                exported_at=exported_at,
            )
        except OSError as exc:
            raise ExportIncompleteError(f"manifest.json 写入失败：{exc}") from exc

        # 6) 全部成功后才标记已导出；状态落盘失败视为批次失败，不宣称成功
        try:
            self.state_store.mark_exported(
                record_ids, batch_id=batch_id, exported_at=exported_at
            )
        except OSError as exc:
            raise ExportIncompleteError(f"导出状态写入失败：{exc}") from exc
        return ExportResult(
            output_dir=output_dir,
            batch_id=batch_id,
            exported_at=exported_at,
            record_ids=record_ids,
            warnings=warnings,
        )

    # ------------------------------------------------------------------ Excel

    @staticmethod
    def _excel_thumbnail(source: Path) -> tuple[bytes, int, int] | None:
        """生成嵌入 Excel 的主图缩略图（最长边 ≤ 112px），不嵌高清原图。

        单张图片损坏时返回 None（降级为不嵌入 + warning），不阻断整批导出。
        """
        try:
            from PIL import Image as PILImage
        except ImportError as exc:  # pragma: no cover - 依赖缺失场景
            raise ExportIncompleteError("缺少 Pillow，无法嵌入主图缩略图") from exc
        try:
            with PILImage.open(source) as image:
                image.thumbnail((_EXCEL_MAIN_IMAGE_MAX_PX, _EXCEL_MAIN_IMAGE_MAX_PX))
                size = image.size
                buffer = io.BytesIO()
                image.convert("RGB").save(buffer, format="JPEG", quality=85)
                return buffer.getvalue(), size[0], size[1]
        except (OSError, ValueError):
            return None

    def _write_excel(
        self,
        path: Path,
        *,
        records: list[dict[str, Any]],
        image_relative_paths: list[list[str]],
        batch_id: str,
        exported_at: str,
    ) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - 依赖缺失场景
            raise ExportIncompleteError("缺少 openpyxl，无法生成 Excel") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "校准反馈"
        sheet.append(list(SHEET1_COLUMNS))

        embed_warnings: list[str] = []
        for index, payload in enumerate(records):
            feedback = self._load_feedback(payload)
            row = [
                index + 1,
                first_ai_short_name(payload) or "—",
                str(payload.get("product_link") or ""),
                "",  # 图片列只嵌入主图缩略图，不再写路径文本
                first_ai_shipment_text(payload),
                user_calibration_text(feedback) if feedback is not None else "",
                actual_first_mile_text(feedback) if feedback is not None else "",
            ]
            sheet.append(row)
            # 只嵌入第一张主图缩略图；无图记录保持空白
            relative_paths = image_relative_paths[index]
            if relative_paths:
                main_source = path.parent / relative_paths[0]
                thumb = self._excel_thumbnail(main_source) if main_source.is_file() else None
                if thumb is None:
                    embed_warnings.append(f"记录 {payload.get('id')} 主图缩略图生成失败，图片列留空")
                else:
                    data, width_px, height_px = thumb
                    xl_image = XLImage(io.BytesIO(data))
                    xl_image.width = width_px
                    xl_image.height = height_px
                    sheet_row = index + 2
                    sheet.add_image(xl_image, f"D{sheet_row}")
                    sheet.row_dimensions[sheet_row].height = _EXCEL_ROW_HEIGHT_PT

        widths = (8, 22, 38, 20, 30, 36, 18)
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(SHEET1_COLUMNS))}{sheet.max_row}"
        top_left = Alignment(vertical="top", wrap_text=True)
        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = header_font
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = top_left

        info = workbook.create_sheet("导出信息")
        info.append(list(SHEET2_COLUMNS))
        for index, payload in enumerate(records):
            info.append(
                [
                    str(payload.get("id") or ""),
                    batch_id,
                    exported_at,
                    "\n".join(image_relative_paths[index]),
                    CONTRACT_VERSION,
                ]
            )
        info.freeze_panes = "A2"
        info.auto_filter.ref = f"A1:{get_column_letter(len(SHEET2_COLUMNS))}{info.max_row}"
        for column, width in enumerate((38, 34, 30, 40, 28), start=1):
            info.column_dimensions[get_column_letter(column)].width = width
        for cell in info[1]:
            cell.font = header_font
        for row in info.iter_rows(min_row=2, max_row=info.max_row):
            for cell in row:
                cell.alignment = top_left
        if embed_warnings:
            # 嵌入失败不阻断导出，但写入 Sheet2 备注便于排查
            info.append([])
            info.append(["embed_warnings", "\n".join(embed_warnings), "", "", ""])

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)

    # ------------------------------------------------------------------ manifest

    def _write_manifest(
        self,
        path: Path,
        *,
        records: list[dict[str, Any]],
        image_relative_paths: list[list[str]],
        batch_id: str,
        exported_at: str,
    ) -> None:
        """manifest.json：供 Agent 直接读取的结构化校准数据。

        字段语义与 Sheet1 7 列一致；每条 record 新增 machine_facts
        （ai_initial / user_feedback 两层机器可读事实）；不含 current_estimate、
        采购成本、总成本、标价、利润、补贴、尾程、汇率等经济字段；图片全部相对路径。
        """
        manifest_records: list[dict[str, Any]] = []
        for index, payload in enumerate(records):
            feedback = self._load_feedback(payload)
            relative_paths = list(image_relative_paths[index])
            manifest_records.append(
                {
                    "sequence": index + 1,
                    "record_id": str(payload.get("id") or ""),
                    "product_short_name": first_ai_short_name(payload) or "—",
                    "product_link": str(payload.get("product_link") or ""),
                    "main_image": relative_paths[0] if relative_paths else "",
                    "images": relative_paths,
                    "ai_initial_shipment": first_ai_shipment_text(payload),
                    "user_calibration": user_calibration_text(feedback) if feedback is not None else "",
                    "actual_first_mile": actual_first_mile_text(feedback) if feedback is not None else "",
                    "machine_facts": {
                        "ai_initial": _machine_ai_initial(payload),
                        "local_adopted": _machine_local_adopted(_ai_initial_block(payload)),
                        "user_feedback": _machine_user_feedback(feedback),
                        "actual_logistics": (
                            _machine_user_feedback(feedback).get("actual_logistics")
                            if feedback is not None
                            else None
                        ),
                    },
                }
            )
        manifest = _strip_forbidden_keys({
            "contract_version": CONTRACT_VERSION,
            "export_batch_id": batch_id,
            "exported_at": exported_at,
            "records": manifest_records,
        })
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        temporary.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        temporary.replace(path)

    def _load_feedback(self, payload: dict[str, Any]):
        v2 = payload.get("_v2") if isinstance(payload.get("_v2"), dict) else {}
        feedback_id = v2.get("calibration_feedback_id")
        if not feedback_id:
            return None
        try:
            return self.feedback_service.load(str(feedback_id))
        except KeyError:
            return None
