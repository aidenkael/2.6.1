"""阶段 3：校准反馈导出 V2 服务测试。

覆盖：
- CONTRACT_VERSION = Calibration Feedback Export V2；
- Excel Sheet1 严格 7 列、Sheet2 技术元数据结构不变；
- 原有人类可读 manifest 字段语义不变，每条 record 新增 machine_facts；
- machine_facts.ai_initial：observation / evidence 白名单过滤、
  packaging_proposal（external 优先、adopted 回退）精确输出；
- machine_facts.user_feedback：structure / suggested_package /
  actual_logistics 精确输出；
- current_estimate 与经济字段递归排除（AI observation、raw evidence 均过滤）；
- legacy / 缺失 AI initial 不报错、不伪造精确 AI 事实；
- 图片复制、thumbnail fallback、pending/range/all、失败事务测试继续通过。
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

pytest.importorskip("PySide6")
pytest.importorskip("PIL")

from PySide6.QtCore import QBuffer  # noqa: E402
from PySide6.QtGui import QColor, QImage  # noqa: E402

from profit_accounting_26.application import AppContext
from profit_accounting_26.application.calibration_export_service import (
    CONTRACT_VERSION,
    CalibrationFeedbackExporter,
    ExportIncompleteError,
    ExportStateStore,
    SHEET1_COLUMNS,
    SHEET2_COLUMNS,
    first_ai_short_name,
    first_ai_shipment_text,
    parse_seq_range,
)

@pytest.fixture()
def context(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("PROFIT_ACCOUNTING_DATA_DIR", str(tmp_path))
    return AppContext.create_default()


def _png_bytes() -> bytes:
    image = QImage(8, 8, QImage.Format.Format_RGB32)
    image.fill(QColor("#336699"))
    buffer = QBuffer()
    buffer.open(QBuffer.OpenModeFlag.WriteOnly)
    image.save(buffer, "PNG")
    return bytes(buffer.data())


def _payload(product_name: str = "测试商品", *, forwarder: str = "深圳货代") -> dict:
    return {
        "product_name": product_name,
        "product_link": f"https://detail.1688.com/offer/{product_name}.html",
        "product_cost_rmb": 66.80,
        "domestic_shipping_rmb": 28.00,
        "layers": {
            "adopted": {
                "selected_packaging": "保守档",
                "normal": {
                    "packaging_method": "气泡袋",
                    "length_cm": 17,
                    "width_cm": 32,
                    "height_cm": 17,
                    "weight_g": 720,
                },
                "conservative": {
                    "packaging_method": "气泡袋",
                    "length_cm": 17,
                    "width_cm": 32,
                    "height_cm": 17,
                    "weight_g": 720,
                },
                "bare": {"length_cm": 45, "width_cm": 30, "height_cm": 15, "weight_g": 580},
            },
            "calculated": {
                "system_cost_rmb": 237.26,
                "exchange_rate": 7.2,
                "forwarder_name": forwarder,
                "logistics_quote": {
                    "weight_fee_rmb": 92.48,
                    "fixed_fee_rmb": 10.0,
                    "tail_fee_rmb": 39.98,
                    "total_logistics_rmb": 142.46,
                },
            },
        },
        "profit_scenarios": {
            "driver": "no_activity_price",
            "no_activity": {
                "sale_price_usd": 30.0,
                "profit_rmb": 78.74,
                "profit_rate_on_cost": 0.5,
                "rule_status": {},
            },
            "activity": {
                "sale_price_usd": 27.0,
                "profit_rmb": 55.12,
                "profit_rate_on_cost": 0.35,
                "rule_status": {},
            },
        },
    }


def _ai_initial(*, product_name: str = "AI首次简名") -> dict:
    return {
        "observation": {
            "product_name": product_name,
            "length_cm": 25,
            "width_cm": 20,
            "height_cm": 5,
            "weight_g": 600,
            "raw_payload": {
                "shipment": {
                    "length_cm": 25,
                    "width_cm": 20,
                    "height_cm": 5,
                    "weight_g": 600,
                    "state": "可压缩；袋装发货",
                }
            },
        },
        "external_ai_packaging_proposal": {
            "normal": {
                "packaging_method": "可压缩；袋装发货",
                "length_cm": 25,
                "width_cm": 20,
                "height_cm": 5,
                "weight_g": 600,
            }
        },
        "adopted_packaging": {
            "normal": {
                "packaging_method": "可压缩；袋装发货",
                "length_cm": 25,
                "width_cm": 20,
                "height_cm": 5,
                "weight_g": 600,
            }
        },
    }


def _create_v2(context, *, product_name: str = "测试商品", ai_initial=None, images=None) -> str:
    payload = _payload(product_name)
    if images is not None:
        payload["images"] = images
    return context.history_record_v2_service.create_record(
        payload,
        ai_initial=ai_initial if ai_initial is not None else _ai_initial(),
    )


def _image_refs(context, count: int = 1) -> list[dict]:
    refs = []
    for index in range(count):
        reference = context.image_store.add_bytes(
            _png_bytes(), suffix=".png", original_filename=f"img{index}.png"
        )
        refs.append(
            {
                **reference.to_dict(),
                "order": index,
                "relative_path": reference.storage_key,
            }
        )
    return refs


def _load_workbook(path: Path):
    from openpyxl import load_workbook

    return load_workbook(path)


class TestExportShape:
    def test_sheet1_exactly_seven_columns(self, context, tmp_path):
        _create_v2(context)
        records = context.record_service.list()
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        sheet = workbook["校准反馈"]
        assert sheet.max_column == 7
        assert [cell.value for cell in sheet[1]] == list(SHEET1_COLUMNS)

    def test_sheet1_has_no_actual_logistics_column(self, context, tmp_path):
        """本轮删除“实际物流数据”列；底层 ActualLogistics 数据结构不变。"""
        _create_v2(context)
        records = context.record_service.list()
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        headers = [cell.value for cell in workbook["校准反馈"][1]]
        assert "实际物流数据" not in headers
        assert headers == ["序号", "商品简名", "商品链接", "图片", "AI首次发货估算", "用户校准内容", "真实头程"]

    def test_sheet2_metadata_and_contract_version(self, context, tmp_path):
        record_id = _create_v2(context)
        records = context.record_service.list()
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        info = workbook["导出信息"]
        assert [cell.value for cell in info[1]] == list(SHEET2_COLUMNS)
        row2 = [cell.value for cell in info[2]]
        assert row2[0] == record_id
        assert row2[1] == result.batch_id
        assert row2[4] == CONTRACT_VERSION

    def test_export_state_file_structure(self, context, tmp_path):
        record_id = _create_v2(context)
        records = context.record_service.list()
        result = context.calibration_export_service.export(records, "all", tmp_path)
        state_path = context.paths.data_dir / "calibration" / "export_state.json"
        assert state_path.is_file()
        store = ExportStateStore(context.paths.data_dir)
        assert record_id in store.exported_record_ids()
        assert result.record_ids == [record_id]


class TestExportSources:
    def test_first_ai_short_name_not_overwritten_by_current_name(self, context, tmp_path):
        _create_v2(context, product_name="用户后来改的名称", ai_initial=_ai_initial(product_name="AI首次简名"))
        records = context.record_service.list()
        assert first_ai_short_name(records[0]) == "AI首次简名"
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        assert workbook["校准反馈"].cell(2, 2).value == "AI首次简名"

    def test_legacy_short_name_empty(self, context, tmp_path):
        payload = _payload("旧商品")
        record_id = context.history_record_v2_service.create_record(
            payload, ai_initial=None, record_id="legacy-export-1"
        )
        records = [context.store.load_record(record_id)]
        assert first_ai_short_name(records[0]) == ""
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        assert workbook["校准反馈"].cell(2, 2).value in ("", "—")

    def test_ai_first_shipment_uses_ai_initial_not_current_estimate(self, context, tmp_path):
        _create_v2(context)
        records = context.record_service.list()
        # 当前采用与 AI 首次不同：导出必须用 AI 首次
        text = first_ai_shipment_text(records[0])
        assert "25×20×5 cm / 600g" in text
        assert "可压缩；袋装发货" in text
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        assert workbook["校准反馈"].cell(2, 5).value == "25×20×5 cm / 600g\n可压缩；袋装发货"

    def test_user_calibration_only_user_note_and_suggested(self, context, tmp_path):
        record_id = _create_v2(context)
        service = context.calibration_feedback_service
        feedback_id = service.save(
            {
                "record_id": record_id,
                "user_note": "正常袋装即可，不需要盒装",
                "suggested_package": {
                    "length_cm": 25, "width_cm": 20, "height_cm": 4, "weight_g": 550,
                },
            }
        )
        context.history_record_v2_service.link_feedback(record_id, feedback_id)
        records = context.record_service.list()
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        cell = workbook["校准反馈"].cell(2, 6).value
        assert "用户反馈：正常袋装即可，不需要盒装" in cell
        assert "建议包装：25×20×4 cm / 550g" in cell

    def test_first_mile_empty_when_no_actual(self, context, tmp_path):
        record_id = _create_v2(context)
        service = context.calibration_feedback_service
        feedback_id = service.save({"record_id": record_id, "user_note": "只有文字反馈"})
        context.history_record_v2_service.link_feedback(record_id, feedback_id)
        records = context.record_service.list()
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        assert workbook["校准反馈"].cell(2, 7).value in (None, "")

    def test_first_mile_uses_actual_forwarder_and_fee(self, context, tmp_path):
        """真实头程语义不变：actual_forwarder + actual_first_mile_fee_rmb（第 7 列）。"""
        record_id = _create_v2(context)
        service = context.calibration_feedback_service
        feedback_id = service.save(
            {
                "record_id": record_id,
                "user_note": "实测",
                "actual_logistics": {
                    "actual_package_dimensions": {"length_cm": 24, "width_cm": 19, "height_cm": 4},
                    "actual_package_weight_g": 530,
                    "actual_chargeable_weight_kg": 0.53,
                    "actual_packaging_method": "袋装",
                    "actual_forwarder": "深圳",
                    "actual_first_mile_fee_rmb": 26.0,
                    "evidence_level": "actual_measured",
                },
            }
        )
        context.history_record_v2_service.link_feedback(record_id, feedback_id)
        records = context.record_service.list()
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        assert workbook["校准反馈"].cell(2, 7).value == "深圳 / ¥26.00"


class TestExportImages:
    def test_all_originals_copied_and_main_image_embedded(self, context, tmp_path):
        """images/ 复制全部原图；Sheet1 只嵌入第一张缩略图，不嵌全部高清原图。"""
        refs = _image_refs(context, count=2)
        record_id = _create_v2(context, images=refs)
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        images_dir = result.output_dir / "images"
        assert (images_dir / "001_1.png").is_file()
        assert (images_dir / "001_2.png").is_file()
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        sheet = workbook["校准反馈"]
        # 图片列不再写路径文本
        assert sheet.cell(2, 4).value in (None, "")
        # Sheet2 仍保留相对路径技术字段
        info = workbook["导出信息"].cell(2, 4).value
        assert "images/001_1.png" in info
        # 只嵌入一张主图缩略图（两张原图不得全部嵌入）
        assert len(sheet._images) == 1
        embedded = sheet._images[0]
        assert embedded.width <= 140 and embedded.height <= 140
        assert str(embedded.anchor._from.row) == "1"  # 锚定在第一条记录行（0 基）

    def test_record_without_images_embeds_nothing(self, context, tmp_path):
        record_id = _create_v2(context, images=[])
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        sheet = workbook["校准反馈"]
        assert sheet.cell(2, 4).value in (None, "")
        assert len(sheet._images) == 0

    def test_thumbnail_fallback_records_warning(self, context, tmp_path):
        ref = _image_refs(context, count=1)[0]
        record_id = _create_v2(context, images=[ref])
        # 删除原图，只保留缩略图
        (context.paths.data_dir / ref["storage_key"]).unlink()
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        assert result.warnings, "缩略图 fallback 必须产生 warning"
        assert (result.output_dir / "images" / "001_1.jpg").is_file()

    def test_missing_original_and_thumbnail_fails_without_mark(self, context, tmp_path):
        ref = _image_refs(context, count=1)[0]
        record_id = _create_v2(context, images=[ref])
        (context.paths.data_dir / ref["storage_key"]).unlink()
        thumbnail = ref.get("thumbnail_key")
        if thumbnail:
            (context.paths.data_dir / thumbnail).unlink(missing_ok=True)
        records = [context.store.load_record(record_id)]
        with pytest.raises(ExportIncompleteError):
            context.calibration_export_service.export(records, "all", tmp_path)
        store = ExportStateStore(context.paths.data_dir)
        assert record_id not in store.exported_record_ids()

    def test_record_without_images_allowed(self, context, tmp_path):
        record_id = _create_v2(context, images=[])
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        assert workbook["校准反馈"].cell(2, 4).value in (None, "")


class TestManifest:
    """manifest.json：供 Agent 直接读取的结构化校准数据。"""

    def _export_with_feedback(self, context, tmp_path, *, images=None):
        refs = _image_refs(context, count=2) if images else []
        record_id = _create_v2(context, images=refs)
        service = context.calibration_feedback_service
        feedback_id = service.save(
            {
                "record_id": record_id,
                "user_note": "实际是压缩袋装",
                "suggested_package": {"length_cm": 25, "width_cm": 20, "height_cm": 4, "weight_g": 550},
                "actual_logistics": {
                    "actual_forwarder": "深圳",
                    "actual_first_mile_fee_rmb": 26.0,
                    "evidence_level": "actual_logistics",
                },
            }
        )
        context.history_record_v2_service.link_feedback(record_id, feedback_id)
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        return record_id, result

    def test_manifest_exists_and_parseable(self, context, tmp_path):
        record_id, result = self._export_with_feedback(context, tmp_path, images=True)
        manifest_path = result.output_dir / "manifest.json"
        assert manifest_path.is_file()
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        assert manifest["contract_version"] == CONTRACT_VERSION
        assert manifest["export_batch_id"] == result.batch_id
        assert manifest["exported_at"] == result.exported_at
        assert len(manifest["records"]) == 1
        entry = manifest["records"][0]
        assert entry["sequence"] == 1
        assert entry["record_id"] == record_id
        assert entry["product_short_name"] == "AI首次简名"
        assert entry["ai_initial_shipment"] == "25×20×5 cm / 600g\n可压缩；袋装发货"
        assert "用户反馈：实际是压缩袋装" in entry["user_calibration"]
        assert entry["actual_first_mile"] == "深圳 / ¥26.00"
        # 人可直接阅读：ensure_ascii=False，中文不是 \uXXXX
        raw = manifest_path.read_text(encoding="utf-8")
        assert "用户反馈" in raw

    def test_manifest_image_relative_paths_exist_on_disk(self, context, tmp_path):
        _record_id, result = self._export_with_feedback(context, tmp_path, images=True)
        manifest = json.loads((result.output_dir / "manifest.json").read_text(encoding="utf-8"))
        entry = manifest["records"][0]
        assert entry["main_image"] == "images/001_1.png"
        assert entry["images"] == ["images/001_1.png", "images/001_2.png"]
        for relative in entry["images"]:
            assert not Path(relative).is_absolute()
            assert (result.output_dir / relative).is_file()

    def test_manifest_has_no_economic_or_current_estimate_fields(self, context, tmp_path):
        _record_id, result = self._export_with_feedback(context, tmp_path, images=True)
        raw = (result.output_dir / "manifest.json").read_text(encoding="utf-8")
        manifest = json.loads(raw)
        allowed_top = {"contract_version", "export_batch_id", "exported_at", "records"}
        assert set(manifest) == allowed_top
        allowed_record = {
            "sequence", "record_id", "product_short_name", "product_link",
            "main_image", "images", "ai_initial_shipment",
            "user_calibration", "actual_first_mile", "machine_facts",
        }
        for entry in manifest["records"]:
            assert set(entry) == allowed_record
        for forbidden in (
            "current_estimate", "product_cost", "domestic_shipping", "system_cost",
            "sale_price", "profit_rmb", "profit_usd", "profit_rate",
            "subsidy", "tail_fee", "exchange_rate", "shein_quote",
        ):
            assert forbidden not in raw

    def test_manifest_without_images_keeps_empty_paths(self, context, tmp_path):
        _record_id, result = self._export_with_feedback(context, tmp_path, images=False)
        manifest = json.loads((result.output_dir / "manifest.json").read_text(encoding="utf-8"))
        entry = manifest["records"][0]
        assert entry["main_image"] == ""
        assert entry["images"] == []

    def test_manifest_write_failure_does_not_mark_exported(self, context, tmp_path, monkeypatch):
        record_id = _create_v2(context)
        records = context.record_service.list()

        def _boom(*_args, **_kwargs):
            raise OSError("模拟 manifest 写入失败")

        monkeypatch.setattr(context.calibration_export_service, "_write_manifest", _boom)
        with pytest.raises(ExportIncompleteError):
            context.calibration_export_service.export(records, "all", tmp_path)
        assert record_id not in ExportStateStore(context.paths.data_dir).exported_record_ids()


class TestManifestV2:
    """V2 合同：machine_facts 机器可读事实 + 经济字段递归排除。"""

    def _export_one(self, context, tmp_path, *, ai_initial=None, images=None) -> tuple[str, Path]:
        refs = _image_refs(context, count=1) if images else []
        record_id = _create_v2(context, ai_initial=ai_initial, images=refs)
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        return record_id, result.output_dir / "manifest.json"

    def _read_manifest(self, manifest_path: Path) -> dict:
        return json.loads(manifest_path.read_text(encoding="utf-8"))

    @staticmethod
    def _walk(value, predicate):
        """递归遍历 manifest，对每个键名执行断言。"""
        if isinstance(value, dict):
            for key, item in value.items():
                predicate(str(key))
                TestManifestV2._walk(item, predicate)
        elif isinstance(value, list):
            for item in value:
                TestManifestV2._walk(item, predicate)

    def test_contract_version_is_v2(self):
        assert CONTRACT_VERSION == "Calibration Feedback Export V2"

    def test_top_level_and_record_keys_plus_machine_facts(self, context, tmp_path):
        _record_id, manifest_path = self._export_one(context, tmp_path, images=True)
        manifest = self._read_manifest(manifest_path)
        assert set(manifest) == {"contract_version", "export_batch_id", "exported_at", "records"}
        entry = manifest["records"][0]
        assert set(entry) == {
            "sequence", "record_id", "product_short_name", "product_link",
            "main_image", "images", "ai_initial_shipment",
            "user_calibration", "actual_first_mile", "machine_facts",
        }
        assert set(entry["machine_facts"]) == {
            "ai_initial", "local_adopted", "user_feedback", "actual_logistics", "reestimate_history",
        }

    def test_machine_facts_ai_initial_observation_exact_values(self, context, tmp_path):
        _record_id, manifest_path = self._export_one(context, tmp_path)
        manifest = self._read_manifest(manifest_path)
        ai_initial = manifest["records"][0]["machine_facts"]["ai_initial"]
        observation = ai_initial["observation"]
        assert observation["product_name"] == "AI首次简名"
        assert observation["length_cm"] == 25
        assert observation["width_cm"] == 20
        assert observation["height_cm"] == 5
        assert observation["weight_g"] == 600
        assert ai_initial["packaging_proposal"]["source_kind"] == "external_ai_packaging_proposal"
        assert ai_initial["packaging_proposal"]["normal"]["length_cm"] == 25

    def test_observation_economic_fields_never_enter_manifest(self, context, tmp_path):
        ai_initial = _ai_initial()
        ai_initial["observation"]["product_cost_rmb"] = 66.8
        ai_initial["observation"]["product_cost_value_type"] = "page_estimate"
        ai_initial["observation"]["domestic_shipping_rmb"] = 28.0
        ai_initial["observation"]["domestic_shipping_value_type"] = "page_estimate"
        _record_id, manifest_path = self._export_one(context, tmp_path, ai_initial=ai_initial)
        raw = manifest_path.read_text(encoding="utf-8")
        observation = self._read_manifest(manifest_path)["records"][0]["machine_facts"]["ai_initial"]["observation"]
        assert "product_cost_rmb" not in observation
        assert "product_cost_value_type" not in observation
        assert "domestic_shipping_rmb" not in observation
        assert "domestic_shipping_value_type" not in observation
        assert "product_cost_rmb" not in raw
        assert "domestic_shipping_rmb" not in raw

    def test_raw_evidence_whitelist_keeps_structure_drops_economics(self, context, tmp_path):
        ai_initial = _ai_initial()
        ai_initial["observation"]["raw_payload"] = {
            "field_evidence": {
                "has_hard_bottom": {"observed": True, "confidence": "high"},
                "material": {"value": "塑料", "confidence": "medium"},
                "product_cost_rmb": {"value": 66.8},
                "domestic_shipping_rmb": {"value": 28.0},
            },
            "confirmed_facts": {
                "length_cm": {"value": 25},
                "profit_rmb": {"value": 100},
            },
            "dimension_semantic_issue": "dimension_evidence_not_outer_dimensions",
            "shipment": {
                "length_cm": 25, "width_cm": 20, "height_cm": 5, "weight_g": 600,
                "state": "可压缩；袋装发货", "packaging_method": "袋装",
                "total_fee_rmb": 999,
            },
        }
        _record_id, manifest_path = self._export_one(context, tmp_path, ai_initial=ai_initial)
        raw = manifest_path.read_text(encoding="utf-8")
        evidence = self._read_manifest(manifest_path)["records"][0]["machine_facts"]["ai_initial"]["evidence"]
        assert evidence["field_evidence"]["has_hard_bottom"] == {"observed": True, "confidence": "high"}
        assert "material" in evidence["field_evidence"]
        assert "product_cost_rmb" not in evidence["field_evidence"]
        assert "domestic_shipping_rmb" not in evidence["field_evidence"]
        assert evidence["confirmed_facts"] == {"length_cm": {"value": 25}}
        assert evidence["dimension_semantic_issue"] == "dimension_evidence_not_outer_dimensions"
        assert evidence["shipment"] == {
            "length_cm": 25, "width_cm": 20, "height_cm": 5, "weight_g": 600,
            "state": "可压缩；袋装发货", "packaging_method": "袋装",
        }
        for forbidden in ("product_cost_rmb", "domestic_shipping_rmb", "profit_rmb", "total_fee_rmb"):
            assert forbidden not in raw

    def test_packaging_proposal_exports_scenario_whitelist_only(self, context, tmp_path):
        ai_initial = _ai_initial()
        ai_initial["external_ai_packaging_proposal"] = {
            "normal": {
                "label": "normal",
                "packaging_state": "moderate_compression",
                "packaging_method": "可压缩；袋装发货",
                "length_cm": 25, "width_cm": 20, "height_cm": 5, "weight_g": 600,
                "reasoning_summary": "软质可压缩",
                "confidence": "medium",
                "needs_review": True,
                "default_fields_used": ["packaging_state"],
            },
            "conservative": {
                "label": "conservative",
                "packaging_method": "盒装",
                "length_cm": 26, "width_cm": 21, "height_cm": 6, "weight_g": 650,
            },
            "proposal_source": "external_ai",
            "engine_version": "packaging-estimation-v1",
            "calibration_version": "local-calibration-v3",
        }
        _record_id, manifest_path = self._export_one(context, tmp_path, ai_initial=ai_initial)
        proposal = self._read_manifest(manifest_path)["records"][0]["machine_facts"]["ai_initial"]["packaging_proposal"]
        assert proposal["source_kind"] == "external_ai_packaging_proposal"
        normal = proposal["normal"]
        assert normal["packaging_method"] == "可压缩；袋装发货"
        assert normal["length_cm"] == 25 and normal["weight_g"] == 600
        assert set(normal) == {
            "packaging_state", "packaging_method", "length_cm", "width_cm", "height_cm",
            "weight_g", "reasoning_summary", "confidence", "needs_review", "default_fields_used",
        }
        assert proposal["conservative"]["length_cm"] == 26
        assert proposal["conservative"]["packaging_method"] == "盒装"
        assert proposal["engine_version"] == "packaging-estimation-v1"
        assert proposal["calibration_version"] == "local-calibration-v3"

    def test_packaging_proposal_falls_back_to_adopted_packaging(self, context, tmp_path):
        ai_initial = _ai_initial()
        ai_initial["external_ai_packaging_proposal"] = None
        _record_id, manifest_path = self._export_one(context, tmp_path, ai_initial=ai_initial)
        proposal = self._read_manifest(manifest_path)["records"][0]["machine_facts"]["ai_initial"]["packaging_proposal"]
        assert proposal["source_kind"] == "ai_initial.adopted_packaging"
        assert proposal["normal"]["length_cm"] == 25
        assert proposal["normal"]["width_cm"] == 20

    def test_user_feedback_structure_exact_output(self, context, tmp_path):
        record_id = _create_v2(context)
        service = context.calibration_feedback_service
        feedback_id = service.save(
            {
                "record_id": record_id,
                "source": "developer",
                "structure": {
                    "can_fold": True,
                    "can_compress": True,
                    "can_coil": False,
                    "can_disassemble": "unknown",
                    "requires_shape_retention": False,
                    "foldable_parts": ["主体"],
                    "compressible_parts": ["主体"],
                    "coilable_parts": [],
                    "detachable_parts": [],
                    "rigid_parts": ["底板"],
                    "axis_behavior": {"length": "compress", "width": "fold", "height": "preserve"},
                },
            }
        )
        context.history_record_v2_service.link_feedback(record_id, feedback_id)
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        manifest = self._read_manifest(result.output_dir / "manifest.json")
        user_feedback = manifest["records"][0]["machine_facts"]["user_feedback"]
        assert user_feedback["feedback_id"] == feedback_id
        assert user_feedback["feedback_schema_version"] == "calibration-feedback-v1"
        assert user_feedback["source"] == "developer"
        assert user_feedback["created_at"]
        assert user_feedback["updated_at"]
        assert user_feedback["structure"] == {
            "can_fold": True,
            "can_compress": True,
            "can_coil": False,
            "can_disassemble": "unknown",
            "requires_shape_retention": False,
            "foldable_parts": ["主体"],
            "compressible_parts": ["主体"],
            "coilable_parts": [],
            "detachable_parts": [],
            "rigid_parts": ["底板"],
            "axis_behavior": {"length": "compress", "width": "fold", "height": "preserve"},
        }

    def test_suggested_package_exact_output_with_user_suggested_evidence(self, context, tmp_path):
        record_id = _create_v2(context)
        service = context.calibration_feedback_service
        feedback_id = service.save(
            {
                "record_id": record_id,
                "suggested_package": {
                    "length_cm": 25, "width_cm": 20, "height_cm": 4, "weight_g": 550,
                    "packaging_method": "压缩袋装",
                    "evidence_level": "actual_measured",  # 合同强制降级为 user_suggested
                },
            }
        )
        context.history_record_v2_service.link_feedback(record_id, feedback_id)
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        manifest = self._read_manifest(result.output_dir / "manifest.json")
        suggested = manifest["records"][0]["machine_facts"]["user_feedback"]["suggested_package"]
        assert suggested == {
            "length_cm": 25.0, "width_cm": 20.0, "height_cm": 4.0, "weight_g": 550.0,
            "packaging_method": "压缩袋装",
            "evidence_level": "user_suggested",
        }

    def test_actual_logistics_exact_output(self, context, tmp_path):
        record_id = _create_v2(context)
        service = context.calibration_feedback_service
        feedback_id = service.save(
            {
                "record_id": record_id,
                "actual_logistics": {
                    "actual_first_mile_fee_rmb": 26.0,
                    "actual_forwarder": "深圳",
                    "actual_chargeable_weight_kg": 0.53,
                    "actual_package_dimensions": {"length_cm": 24, "width_cm": 19, "height_cm": 4},
                    "actual_package_weight_g": 530,
                    "actual_packaging_method": "袋装",
                    "evidence_level": "actual_measured",
                },
            }
        )
        context.history_record_v2_service.link_feedback(record_id, feedback_id)
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        manifest = self._read_manifest(result.output_dir / "manifest.json")
        user_feedback = manifest["records"][0]["machine_facts"]["user_feedback"]
        actual = user_feedback["actual_logistics"]
        assert actual == {
            "actual_first_mile_fee_rmb": 26.0,
            "actual_forwarder": "深圳",
            "actual_chargeable_weight_kg": 0.53,
            "actual_package_dimensions": {"length_cm": 24, "width_cm": 19, "height_cm": 4},
            "actual_package_weight_g": 530.0,
            "actual_packaging_method": "袋装",
            "evidence_level": "actual_measured",
        }
        # 软件内部导出状态不属于规则分析事实
        for excluded in ("calibration_exported_at", "calibration_export_batch_id",
                         "feedback_updated_after_export", "record_id"):
            assert excluded not in user_feedback

    def test_manifest_still_excludes_current_estimate_recursively(self, context, tmp_path):
        _record_id, manifest_path = self._export_one(context, tmp_path, images=True)
        raw = manifest_path.read_text(encoding="utf-8")
        manifest = self._read_manifest(manifest_path)
        assert "current_estimate" not in raw

        def assert_not_current_estimate(key: str):
            assert key != "current_estimate"

        self._walk(manifest, assert_not_current_estimate)

    def test_manifest_recursive_economic_field_scan(self, context, tmp_path):
        ai_initial = _ai_initial()
        # 误落位经济字段：顶层、observation、raw evidence 三层都塞入
        ai_initial["product_cost_rmb"] = 66.8
        ai_initial["observation"]["sale_price_usd"] = 30.0
        ai_initial["observation"]["raw_payload"] = {
            "field_evidence": {
                "has_hard_bottom": {"observed": True},
                "subsidy": {"value": 5.0},
                "exchange_rate": {"value": 7.2},
            },
            "shipment": {"length_cm": 25, "tail_fee_rmb": 39.98},
        }
        record_id = _create_v2(context, ai_initial=ai_initial)
        service = context.calibration_feedback_service
        feedback_id = service.save({"record_id": record_id, "user_note": "备注"})
        context.history_record_v2_service.link_feedback(record_id, feedback_id)
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        raw = (result.output_dir / "manifest.json").read_text(encoding="utf-8")
        manifest = self._read_manifest(result.output_dir / "manifest.json")
        forbidden_tokens = (
            "current_estimate", "calculation_snapshot", "profit_scenarios",
            "product_cost_rmb", "product_cost_value_type",
            "domestic_shipping_rmb", "domestic_shipping_value_type",
            "system_cost_rmb", "exchange_rate",
            "sale_price_usd", "sale_price_rmb",
            "profit_rmb", "profit_usd", "profit_rate",
            "subsidy", "tail_fee_rmb", "shein_quote_usd",
        )
        for token in forbidden_tokens:
            assert token not in raw
        forbidden_key_parts = (
            "current_estimate", "product_cost", "domestic_shipping", "system_cost",
            "sale_price", "profit", "subsidy", "tail_fee", "exchange_rate", "shein_quote",
        )

        def assert_not_forbidden_key(key: str):
            assert not any(part in key for part in forbidden_key_parts), key

        self._walk(manifest, assert_not_forbidden_key)

    def test_legacy_missing_ai_initial_safe_and_not_fabricated(self, context, tmp_path):
        payload = _payload("旧商品")
        record_id = context.history_record_v2_service.create_record(
            payload, ai_initial=None, record_id="legacy-v2-manifest"
        )
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        manifest = self._read_manifest(result.output_dir / "manifest.json")
        entry = manifest["records"][0]
        assert entry["record_id"] == record_id
        assert entry["product_short_name"] == "—"
        assert entry["machine_facts"]["ai_initial"] is None
        assert entry["machine_facts"]["user_feedback"] is None

    def test_legacy_layers_ai_raw_not_masqueraded_as_first_ai_facts(self, context, tmp_path):
        payload = _payload("旧商品")
        payload["_v2"] = {
            "record_schema_version": "2.6.1",
            "ai_initial": {
                "legacy_layers_ai_raw": {"shipment": {"length_cm": 99, "weight_g": 999}},
            },
        }
        payload["id"] = "legacy-raw-v2-manifest"
        context.store.save_new_record(payload)
        records = [context.store.load_record("legacy-raw-v2-manifest")]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        raw = (result.output_dir / "manifest.json").read_text(encoding="utf-8")
        manifest = self._read_manifest(result.output_dir / "manifest.json")
        assert manifest["records"][0]["machine_facts"]["ai_initial"] is None
        assert "legacy_layers_ai_raw" not in raw

    def test_no_image_record_still_exports_with_machine_facts(self, context, tmp_path):
        _record_id, manifest_path = self._export_one(context, tmp_path, images=False)
        manifest = self._read_manifest(manifest_path)
        entry = manifest["records"][0]
        assert entry["main_image"] == ""
        assert entry["images"] == []
        assert set(entry["machine_facts"]) == {
            "ai_initial", "local_adopted", "user_feedback", "actual_logistics", "reestimate_history",
        }


class TestExportModes:
    def _three_records(self, context) -> list[str]:
        ids = []
        for name in ("商品A", "商品B", "商品C"):
            ids.append(_create_v2(context, product_name=name))
        return ids

    def test_all_mode_allows_reexport(self, context, tmp_path):
        ids = self._three_records(context)
        records = context.record_service.list()
        first = context.calibration_export_service.export(records, "all", tmp_path / "a")
        second = context.calibration_export_service.export(records, "all", tmp_path / "b")
        assert len(first.record_ids) == 3 and len(second.record_ids) == 3
        assert all(record_id in ExportStateStore(context.paths.data_dir).exported_record_ids() for record_id in ids)

    def test_range_mode_maps_display_seq_to_record_id(self, context, tmp_path):
        ids = self._three_records(context)
        records = context.record_service.list()
        result = context.calibration_export_service.export(
            records, "range", tmp_path, seq_range="2-3"
        )
        # 显示序号 2、3 → 对应列表第 2、3 条
        assert result.record_ids == [records[1]["id"], records[2]["id"]]

    def test_range_parse_and_validation(self):
        assert parse_seq_range("1-30", 30) == (1, 30)
        assert parse_seq_range("3-5", 10) == (3, 5)
        for bad in ("", "abc", "5-3", "0-3", "1-99", "1 2"):
            with pytest.raises(ValueError):
                parse_seq_range(bad, 30)

    def test_pending_mode_filters_exported(self, context, tmp_path):
        ids = self._three_records(context)
        records = context.record_service.list()
        first = context.calibration_export_service.export(records, "range", tmp_path / "a", seq_range="1-1")
        assert first.record_ids == [records[0]["id"]]
        pending = context.calibration_export_service.export(records, "pending", tmp_path / "b")
        assert records[0]["id"] not in pending.record_ids
        assert set(pending.record_ids) == {records[1]["id"], records[2]["id"]}


class TestFailureAndLifecycle:
    def test_deleted_record_stale_state_does_not_crash(self, context, tmp_path):
        record_id = _create_v2(context, product_name="待删除")
        records = context.record_service.list()
        context.calibration_export_service.export(records, "all", tmp_path / "a")
        context.record_service.delete_record(record_id)
        other_id = _create_v2(context, product_name="保留商品")
        records = context.record_service.list()
        # 旧状态里有已删除的 record_id，不影响后续导出
        result = context.calibration_export_service.export(records, "pending", tmp_path / "b")
        assert other_id in result.record_ids

    def test_excel_write_failure_does_not_mark(self, context, tmp_path, monkeypatch):
        record_id = _create_v2(context)
        records = context.record_service.list()

        def _boom(*_args, **_kwargs):
            raise OSError("模拟 Excel 写入失败")

        monkeypatch.setattr(
            context.calibration_export_service,
            "_write_excel",
            _boom,
        )
        with pytest.raises(ExportIncompleteError):
            context.calibration_export_service.export(records, "all", tmp_path)
        assert record_id not in ExportStateStore(context.paths.data_dir).exported_record_ids()

    def test_state_write_failure_is_safe_failure(self, context, tmp_path, monkeypatch):
        """导出状态落盘失败：export() 明确失败、无 record 被标记、无未处理异常。"""
        record_id = _create_v2(context)
        records = context.record_service.list()

        def _boom(*_args, **_kwargs):
            raise OSError("模拟导出状态写入失败")

        monkeypatch.setattr(
            context.calibration_export_service.state_store,
            "mark_exported",
            _boom,
        )
        with pytest.raises(ExportIncompleteError):
            context.calibration_export_service.export(records, "all", tmp_path)
        assert record_id not in ExportStateStore(context.paths.data_dir).exported_record_ids()

    def test_legacy_record_export_compatible(self, context, tmp_path):
        payload = _payload("旧商品")
        record_id = context.history_record_v2_service.create_record(
            payload, ai_initial=None, record_id="legacy-export-2"
        )
        records = [context.store.load_record(record_id)]
        result = context.calibration_export_service.export(records, "all", tmp_path)
        workbook = _load_workbook(result.output_dir / "校准反馈.xlsx")
        assert workbook["校准反馈"].max_column == 7
        assert workbook["校准反馈"].cell(2, 2).value in ("", "—")
        assert workbook["校准反馈"].cell(2, 5).value in (None, "")
        assert (result.output_dir / "manifest.json").is_file()

    def test_exported_at_and_batch_id_written_per_record(self, context, tmp_path):
        record_id = _create_v2(context)
        records = context.record_service.list()
        result = context.calibration_export_service.export(records, "all", tmp_path)
        state = ExportStateStore(context.paths.data_dir).load()["records"]
        assert state[record_id]["export_batch_id"] == result.batch_id
        assert state[record_id]["exported_at"] == result.exported_at
