"""手工包装校准导出 Bug 修复 targeted tests（任务书七~十节）。

用户真实反馈：未点 AI 识图、直接在“当前采用”手工填写 长4 宽6 高7 重5g，
底层 machine_facts.user_feedback.suggested_package 保存正确，
但 Excel“用户校准内容”为空、manifest.user_calibration 为空——
根因：is_ai_reestimate_polluted_suggestion() 把
“suggested_package == current_estimate”单独当作 AI 污染证据，
误杀真实用户输入（用户手工修改后 current_estimate 本来就等于 suggested）。

修复原则（任务书九）：只有存在明确 _v2.reestimate_history 且 suggested
与某条 adopted_reestimate_proposal 精确一致，才认为有旧 AI reestimate
污染嫌疑；禁止仅凭 current_estimate 等值隐藏用户建议。

覆盖：
- 纯函数回归：current_estimate 等值不再触发过滤；reestimate_history 精确
  匹配仍过滤；无证据保留；
- 场景 A：无 AI → 手工填包装 → 导出 Excel/manifest 用户校准内容正确，
  machine_facts.user_feedback.suggested_package 原结构保留；
- 场景 B：有 AI initial → 用户手工修改 → 用户修改保留；
- 场景 C：旧版本 reestimate 污染（明确 history + 精确一致）→ 仍执行兼容过滤。
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

pytest.importorskip("PySide6")
pytest.importorskip("PIL")

from profit_accounting_26.application import AppContext  # noqa: E402
from profit_accounting_26.application.calibration_export_service import (  # noqa: E402
    is_ai_reestimate_polluted_suggestion,
)


@pytest.fixture()
def context(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("PROFIT_ACCOUNTING_DATA_DIR", str(tmp_path))
    return AppContext.create_default()


def _payload(product_name: str = "手工记录") -> dict:
    return {
        "product_name": product_name,
        "product_link": f"https://detail.1688.com/offer/{product_name}.html",
        "product_cost_rmb": 10.0,
        "domestic_shipping_rmb": 0.0,
        "layers": {
            "adopted": {
                "selected_packaging": "保守档",
                "normal": {"length_cm": 25, "width_cm": 20, "height_cm": 5, "weight_g": 600},
                "conservative": {"length_cm": 25, "width_cm": 20, "height_cm": 5, "weight_g": 600},
                "bare": {},
            },
            "calculated": {
                "system_cost_rmb": 10.0,
                "exchange_rate": 7.2,
                "forwarder_name": "深圳货代",
                "logistics_quote": {},
            },
        },
    }


def _ai_initial() -> dict:
    return {
        "observation": {
            "product_name": "AI首次简名",
            "length_cm": 25,
            "width_cm": 20,
            "height_cm": 5,
            "weight_g": 600,
            "raw_payload": {
                "shipment": {
                    "length_cm": 25, "width_cm": 20, "height_cm": 5, "weight_g": 600,
                    "state": "可压缩；袋装发货",
                }
            },
        },
        "external_ai_packaging_proposal": {
            "normal": {"length_cm": 25, "width_cm": 20, "height_cm": 5, "weight_g": 600},
        },
        "adopted_packaging": {
            "normal": {"length_cm": 25, "width_cm": 20, "height_cm": 5, "weight_g": 600},
        },
    }


_MANUAL = {"length_cm": 4, "width_cm": 6, "height_cm": 7, "weight_g": 5}


def _create_record(
    context,
    *,
    with_ai: bool,
    current_estimate: dict,
    reestimate_history: list | None = None,
) -> str:
    record_id = context.history_record_v2_service.create_record(
        _payload(),
        ai_initial=_ai_initial() if with_ai else None,
        current_estimate=current_estimate,
    )
    if reestimate_history:
        context.history_record_v2_service.update_record(
            record_id, {"_v2": {"reestimate_history": reestimate_history}},
        )
    return record_id


def _link_manual_feedback(context, record_id: str) -> None:
    feedback_id = context.calibration_feedback_service.save(
        {"record_id": record_id, "suggested_package": dict(_MANUAL)}
    )
    context.history_record_v2_service.link_feedback(record_id, feedback_id)


def _export(context, tmp_path):
    records = context.record_service.list()
    return context.calibration_export_service.export(records, "all", tmp_path)


def _excel_user_calibration(result) -> str:
    from openpyxl import load_workbook
    workbook = load_workbook(result.output_dir / "校准反馈.xlsx")
    return workbook["校准反馈"].cell(2, 6).value or ""


def _manifest(result) -> dict:
    return json.loads((result.output_dir / "manifest.json").read_text(encoding="utf-8"))


class _FakeSuggested:
    """模拟 feedback.suggested_package（与 UserSuggestedPackage 同字段）。"""

    def __init__(self, dims: dict) -> None:
        self.length_cm = dims["length_cm"]
        self.width_cm = dims["width_cm"]
        self.height_cm = dims["height_cm"]
        self.weight_g = dims["weight_g"]
        self.packaging_method = None


class TestPollutionPredicateCore:
    """核心回归：current_estimate 等值不再单独触发过滤。"""

    def test_current_estimate_equality_alone_does_not_filter(self):
        """suggested == current_estimate 但无 reestimate_history → 保留用户输入。"""
        payload = {"_v2": {"current_estimate": dict(_MANUAL)}}
        assert is_ai_reestimate_polluted_suggestion(payload, _FakeSuggested(_MANUAL)) is False

    def test_no_v2_block_keeps_user_input(self):
        assert is_ai_reestimate_polluted_suggestion({}, _FakeSuggested(_MANUAL)) is False

    def test_reestimate_history_exact_match_still_filters(self):
        """场景 C 兼容：明确 history + suggested 与 adopted_reestimate_proposal 精确一致。"""
        payload = {
            "_v2": {
                "current_estimate": dict(_MANUAL),
                "reestimate_history": [
                    {
                        "reestimate_id": "r1",
                        "sequence": 1,
                        "timestamp": "2025-01-01T00:00:00",
                        "accepted": True,
                        "user_correction": "",
                        "adopted_reestimate_proposal": {
                            "normal": dict(_MANUAL),
                        },
                    }
                ],
            }
        }
        assert is_ai_reestimate_polluted_suggestion(payload, _FakeSuggested(_MANUAL)) is True

    def test_reestimate_history_no_match_keeps_user_input(self):
        """history 存在但 suggested 与 adopted 不一致 → 保留用户输入。"""
        payload = {
            "_v2": {
                "current_estimate": dict(_MANUAL),
                "reestimate_history": [
                    {
                        "adopted_reestimate_proposal": {
                            "normal": {"length_cm": 9, "width_cm": 9, "height_cm": 9, "weight_g": 9},
                        }
                    }
                ],
            }
        }
        assert is_ai_reestimate_polluted_suggestion(payload, _FakeSuggested(_MANUAL)) is False


class TestExportScenarioA:
    """场景 A：无 AI → 用户直接填包装 → 保存 → 导出。"""

    def test_excel_and_manifest_show_user_calibration(self, context, tmp_path):
        record_id = _create_record(context, with_ai=False, current_estimate=dict(_MANUAL))
        _link_manual_feedback(context, record_id)
        result = _export(context, tmp_path)

        excel_text = _excel_user_calibration(result)
        assert "建议包装：4×6×7 cm / 5g" in excel_text, f"Excel 用户校准内容应为空? 实际={excel_text!r}"

        manifest = _manifest(result)
        record = manifest["records"][0]
        assert "建议包装：4×6×7 cm / 5g" in record["user_calibration"], (
            f"manifest.user_calibration 应为空? 实际={record['user_calibration']!r}"
        )
        user_feedback = record["machine_facts"]["user_feedback"]
        assert user_feedback is not None
        suggested = user_feedback["suggested_package"]
        assert suggested is not None
        assert suggested["length_cm"] == 4
        assert suggested["width_cm"] == 6
        assert suggested["height_cm"] == 7
        assert suggested["weight_g"] == 5


class TestExportScenarioB:
    """场景 B：有 AI initial → 用户手工修改 → 用户修改必须保留。"""

    def test_user_edit_kept_with_ai_initial(self, context, tmp_path):
        record_id = _create_record(context, with_ai=True, current_estimate=dict(_MANUAL))
        _link_manual_feedback(context, record_id)
        result = _export(context, tmp_path)

        excel_text = _excel_user_calibration(result)
        assert "建议包装：4×6×7 cm / 5g" in excel_text, f"用户修改应保留，实际={excel_text!r}"

        manifest = _manifest(result)
        record = manifest["records"][0]
        # AI 首次估算仍保留原始 AI 值（25×20×5/600g），未被用户修改覆盖
        assert "25×20×5 cm / 600g" in record["ai_initial_shipment"]
        assert "建议包装：4×6×7 cm / 5g" in record["user_calibration"]
        ai_block = record["machine_facts"]["ai_initial"]
        assert ai_block is not None, "有 AI initial 时 machine_facts.ai_initial 不应为空"


class TestExportScenarioC:
    """场景 C：真实旧版本 AI reestimate 污染记录 → 仍执行旧兼容过滤。"""

    def test_polluted_suggestion_still_filtered(self, context, tmp_path):
        polluted_history = [
            {
                "reestimate_id": "r-old-1",
                "sequence": 1,
                "timestamp": "2025-01-01T00:00:00",
                "accepted": True,
                "user_correction": "",
                "adopted_reestimate_proposal": {"normal": dict(_MANUAL)},
            }
        ]
        record_id = _create_record(
            context, with_ai=True, current_estimate=dict(_MANUAL),
            reestimate_history=polluted_history,
        )
        _link_manual_feedback(context, record_id)
        result = _export(context, tmp_path)

        excel_text = _excel_user_calibration(result)
        assert "建议包装" not in excel_text, (
            f"旧 AI reestimate 污染建议应被过滤，实际={excel_text!r}"
        )
        manifest = _manifest(result)
        assert "建议包装" not in manifest["records"][0]["user_calibration"]
        # 底层 machine_facts.user_feedback.suggested_package 结构保留不变（只影响显示/导出）
        user_feedback = manifest["records"][0]["machine_facts"]["user_feedback"]
        assert user_feedback["suggested_package"]["length_cm"] == 4
